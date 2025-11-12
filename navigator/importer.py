"""Import utilities for JSON and CSV/XLSX sources."""

from __future__ import annotations

import csv
import io
import json
from dataclasses import replace
from typing import Dict, Iterable, List, Sequence, Tuple

from .catalog import ensure_unique_ids, merge_catalog, slugify
from .models import Catalog, Qualification, QualificationStatus

try:  # optional dependency for excel files
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover - optional dependency
    openpyxl = None


class ImportResult(dict):
    """Typed dict used by tests and UI."""


def load_qualifications_from_json(stream: io.TextIOBase) -> List[Qualification]:
    payload = json.load(stream)
    qualifications = []
    for item in payload:
        qualifications.append(
            Qualification(
                id=item.get("id", ""),
                title=item["title"],
                category=item.get("category", ""),
                tags=list(item.get("tags", [])),
                status=QualificationStatus(item.get("status", "draft")),
                aliases=list(item.get("aliases", [])),
            )
        )
    return list(ensure_unique_ids(qualifications).values())


def load_tabular_courses(stream: io.BytesIO, *, mapping: Dict[str, str], filetype: str) -> List[Dict[str, str]]:
    """Load courses from CSV or XLSX streams.

    ``mapping`` maps column names to canonical keys (title/category/description).
    """

    if filetype == "csv":
        decoded = io.TextIOWrapper(stream, encoding="utf-8")
        reader = csv.DictReader(decoded)
        rows = [row for row in reader]
    elif filetype == "xlsx":
        if openpyxl is None:
            raise RuntimeError("openpyxl not installed")
        workbook = openpyxl.load_workbook(stream)
        sheet = workbook.active
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            rows.append({header: value for header, value in zip(headers, row)})
    else:
        raise ValueError(f"Unsupported filetype: {filetype}")

    result: List[Dict[str, str]] = []
    for row in rows:
        record = {canonical: row.get(column, "") for column, canonical in mapping.items()}
        result.append(record)
    return result


def deduplicate_import(base: Catalog, imported: Iterable[Qualification]) -> ImportResult:
    """Merge imported qualifications into the base catalogue.

    Duplicates are detected by title + category. New qualifications get a slug
    based id. Existing ones are returned in ``updated``.
    """

    existing = {(q.title.lower(), q.category.lower()): q for q in base.qualifications}
    new_items: List[Qualification] = []
    updated: List[Qualification] = []

    for item in imported:
        key = (item.title.lower(), item.category.lower())
        if key in existing:
            current = existing[key]
            updated.append(
                replace(
                    current,
                    tags=sorted(set(current.tags) | set(item.tags)),
                    aliases=sorted(set(current.aliases) | set(item.aliases)),
                )
            )
        else:
            identifier = item.id or slugify(item.title)
            new_items.append(replace(item, id=identifier))

    catalog = merge_catalog(base, updated + new_items)
    return ImportResult(
        catalog=catalog,
        created=[item.id for item in new_items],
        updated=[item.id for item in updated],
    )

